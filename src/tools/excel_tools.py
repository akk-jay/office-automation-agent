"""Excel 读写工具 — 基于 openpyxl

分层设计：
  - 前三个函数 (read_excel / write_excel / summarize_excel) 是纯 Python 底层实现，
    返回结构化 dict，不依赖任何 AI 框架。
  - 后三个函数 (tool_read_excel / tool_write_excel / tool_summarize_excel)
    是 LangChain @tool 包装版本，供 LLM Agent 调用。
"""

import openpyxl
from pathlib import Path


def read_excel(file_path: str, sheet_name: str = None) -> dict:
    """读取 Excel 文件并返回所有数据行。

    Args:
        file_path: Excel 文件的路径
        sheet_name: 要读取的工作表名称，默认读取第一个工作表

    Returns:
        dict: {"success": bool, "data": list[dict], "row_count": int, "columns": list, "sheet_name": str, "error": str}
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {
                "success": False,
                "data": [],
                "row_count": 0,
                "columns": [],
                "sheet_name": None,
                "error": "文件为空",
            }

        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
        data = []
        for row in rows[1:]:
            record = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    record[headers[i]] = value
            data.append(record)

        return {
            "success": True,
            "data": data,
            "row_count": len(data),
            "columns": headers,
            "sheet_name": ws.title,
            "error": None,
        }
    except FileNotFoundError:
        return {
            "success": False,
            "data": [],
            "row_count": 0,
            "columns": [],
            "sheet_name": None,
            "error": f"文件不存在: {file_path}",
        }
    except Exception as e:
        return {
            "success": False,
            "data": [],
            "row_count": 0,
            "columns": [],
            "sheet_name": None,
            "error": str(e),
        }


def write_excel(data: list, file_path: str, sheet_name: str = "Sheet1") -> dict:
    """将数据列表写入 Excel 文件。

    Args:
        data: 数据列表，每条记录是一个 dict，key 为列名
        file_path: 输出文件路径
        sheet_name: 工作表名称

    Returns:
        dict: {"success": bool, "file_path": str, "row_count": int, "error": str}
    """
    try:
        if not data:
            return {
                "success": False,
                "file_path": file_path,
                "row_count": 0,
                "error": "没有数据可写入",
            }

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # 写表头
        headers = list(data[0].keys())
        ws.append(headers)

        # 写数据行
        for record in data:
            row = [record.get(h) for h in headers]
            ws.append(row)

        Path(file_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(file_path)

        return {
            "success": True,
            "file_path": str(Path(file_path).absolute()),
            "row_count": len(data),
            "error": None,
        }
    except Exception as e:
        return {
            "success": False,
            "file_path": file_path,
            "row_count": 0,
            "error": str(e),
        }


def summarize_excel(file_path: str) -> dict:
    """获取 Excel 文件的概览信息：行数、列名、基本数值统计。

    Args:
        file_path: Excel 文件路径

    Returns:
        dict: {"success": bool, "row_count": int, "columns": list, "sheet_name": str,
               "numeric_stats": dict, "error": str}
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {
                "success": False,
                "row_count": 0,
                "columns": [],
                "sheet_name": None,
                "numeric_stats": {},
                "error": "文件为空",
            }

        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
        data_rows = rows[1:]

        # 统计数值列
        numeric_stats = {}
        for col_idx, header in enumerate(headers):
            values = [
                row[col_idx]
                for row in data_rows
                if col_idx < len(row) and isinstance(row[col_idx], (int, float))
            ]
            if values:
                numeric_stats[header] = {
                    "count": len(values),
                    "sum": sum(values),
                    "avg": round(sum(values) / len(values), 2),
                    "min": min(values),
                    "max": max(values),
                }

        return {
            "success": True,
            "row_count": len(data_rows),
            "columns": headers,
            "sheet_name": ws.title,
            "numeric_stats": numeric_stats,
            "error": None,
        }
    except FileNotFoundError:
        return {
            "success": False,
            "row_count": 0,
            "columns": [],
            "sheet_name": None,
            "numeric_stats": {},
            "error": f"文件不存在: {file_path}",
        }
    except Exception as e:
        return {
            "success": False,
            "row_count": 0,
            "columns": [],
            "sheet_name": None,
            "numeric_stats": {},
            "error": str(e),
        }


# ============================================================
# LangChain Tool 包装版本（供 Agent 使用）
# ============================================================

from langchain_core.tools import tool


@tool
def tool_read_excel(file_path: str, sheet_name: str = "") -> str:
    """读取 Excel 文件并返回所有数据行。

    当用户需要查看、分析或汇总 Excel 表格数据时使用此工具。
    适用场景：用户提到"看下销售数据"、"汇总报表"、"这个 Excel 里有什么"等。

    Args:
        file_path: Excel 文件的完整路径
        sheet_name: 要读取的工作表名称，留空则读取第一个工作表
    """
    result = read_excel(file_path, sheet_name if sheet_name else None)
    if result["success"]:
        lines = [f"工作表: {result['sheet_name']}，共 {result['row_count']} 行"]
        lines.append(f"列名: {', '.join(result['columns'])}")
        lines.append("---数据预览（前 20 行）---")
        for i, row in enumerate(result["data"][:20]):
            lines.append(str(row))
        if result["row_count"] > 20:
            lines.append(f"... 还有 {result['row_count'] - 20} 行未显示")
        return "\n".join(lines)
    return f"读取失败: {result['error']}"


@tool
def tool_write_excel(data_json: str, file_path: str, sheet_name: str = "Sheet1") -> str:
    """将数据写入 Excel 文件。

    当用户需要保存、导出、或生成 Excel 报表时使用此工具。
    适用场景："生成周报"、"导出数据"、"保存为 Excel"等。

    Args:
        data_json: JSON 格式的数据，如 '[{"姓名":"张三","分数":95}]'
        file_path: 输出文件的完整路径
        sheet_name: 工作表名称，默认 Sheet1
    """
    import json
    try:
        data = json.loads(data_json)
    except json.JSONDecodeError:
        return "写入失败: 数据格式不正确，需要 JSON 格式"

    result = write_excel(data, file_path, sheet_name)
    if result["success"]:
        return f"已成功写入 {result['row_count']} 行数据到 {result['file_path']}"
    return f"写入失败: {result['error']}"


@tool
def tool_summarize_excel(file_path: str) -> str:
    """获取 Excel 文件的概览信息（行数、列名、数值统计）。

    当用户想要快速了解 Excel 文件的结构和内容概况时使用。
    适用场景："这个表有多少行"、"销售额大概是多少"、"有哪些列"等。

    Args:
        file_path: Excel 文件的完整路径
    """
    result = summarize_excel(file_path)
    if result["success"]:
        lines = [
            f"文件概览: {result['sheet_name']}",
            f"数据行数: {result['row_count']}",
            f"列名: {', '.join(result['columns'])}",
        ]
        if result.get("numeric_stats"):
            lines.append("数值列统计:")
            for col, stats in result["numeric_stats"].items():
                lines.append(
                    f"  {col}: 总计={stats['sum']}, 平均={stats['avg']}, "
                    f"最大={stats['max']}, 最小={stats['min']}"
                )
        return "\n".join(lines)
    return f"概览失败: {result['error']}"
