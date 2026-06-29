"""Excel 工具函数测试"""
import pytest
import openpyxl
from pathlib import Path


@pytest.fixture
def sample_excel(tmp_path):
    """创建测试用的临时 Excel 文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["日期", "产品", "销量", "单价", "销售额"])
    ws.append(["2026-06-16", "智能门锁 Pro", 12, 2999, 35988])
    ws.append(["2026-06-17", "智能门锁 Lite", 8, 1599, 12792])
    filepath = tmp_path / "test.xlsx"
    wb.save(filepath)
    return filepath


def test_read_excel_returns_all_rows(sample_excel):
    """read_excel 应该返回 Excel 中的所有数据行（不含表头）"""
    from src.tools.excel_tools import read_excel

    result = read_excel(str(sample_excel))

    assert result["success"] is True
    assert len(result["data"]) == 2
    assert result["data"][0]["日期"] == "2026-06-16"
    assert result["data"][0]["销量"] == 12


def test_summarize_excel_returns_stats(sample_excel):
    """summarize_excel 应该返回行数、列名、数值统计"""
    from src.tools.excel_tools import summarize_excel

    result = summarize_excel(str(sample_excel))

    assert result["success"] is True
    assert result["row_count"] == 2
    assert "日期" in result["columns"]
    assert result["sheet_name"] == "Sheet1"
    # 销量和销售额应该有数值统计
    assert "销量" in result["numeric_stats"]
    assert "销售额" in result["numeric_stats"]


def test_write_excel_creates_file(tmp_path):
    """write_excel 应该能创建新 Excel 文件"""
    from src.tools.excel_tools import write_excel

    outpath = tmp_path / "output.xlsx"
    data = [
        {"姓名": "张三", "分数": 95},
        {"姓名": "李四", "分数": 87},
    ]
    result = write_excel(data, str(outpath))

    assert result["success"] is True
    assert Path(outpath).exists()

    # 验证写入内容
    wb = openpyxl.load_workbook(outpath)
    ws = wb.active
    assert ws["A1"].value == "姓名"
    assert ws["A2"].value == "张三"
    assert ws["B2"].value == 95


def test_read_excel_file_not_found():
    """read_excel 对不存在的文件应返回失败"""
    from src.tools.excel_tools import read_excel

    result = read_excel("不存在的文件.xlsx")
    assert result["success"] is False
    assert "不存在" in result["error"]
